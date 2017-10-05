import os
import re

import openpyxl

from Lessons import *


class TimetableParser:

    PATTERN_SPECIFIC_WEEKS = "([0-9, ]+) н ([А-Яа-я \-A-Za-z]+)"
    PATTERN_EXCEPT_WEEKS = "кр ([0-9, ]+) н ([А-Яа-я \-A-Za-z]+)"
    PATTERN_PRACTICE = "(практ|пр)"
    PATTERN_LECTURE = "лек"
    PATTERN_LAB = "лаб"

    FILE_PATH = os.getcwd() + "/files/official_timetable.xlsx"

    def __init__(self, group_name):
        workbook = openpyxl.load_workbook(TimetableParser.FILE_PATH)
        self.timetable_sheet = workbook["Лист1"]
        self.group_name = group_name
        self.timetable = []

    def parse(self):
        self.get_timetable()
        self.add_non_academic_days()

    def get_group_column(self):
        row = 2
        column = 6
        cnt = 0
        found = False

        cell_value = self.timetable_sheet.cell(row=row, column=column).value
        while cell_value:
            if self.group_name in cell_value:
                found = True
                break

            cnt += 1
            column += 4 if cnt % 3 != 0 else 10
            cell_value = self.timetable_sheet.cell(row=row, column=column).value

        return column if found else None

    def get_timetable(self):
        self.init_timetable()

        week_count = utils.get_week_count()
        first_academic_day = utils.get_first_academic_day()
        last_academic_day = utils.get_last_academic_day()
        first_academic_day_index = first_academic_day.weekday()
        last_academic_day_index = week_count * 6 + last_academic_day.weekday()

        col = self.get_group_column()
        discipline_col = col
        type_col = col + 1
        lecturer_col = col + 2
        room_col = col + 3

        for row in range(4, 76):
            discipline = self.timetable_sheet.cell(row=row, column=discipline_col).value

            if not discipline:
                continue

            tp = self.timetable_sheet.cell(row=row, column=type_col).value
            lecturer = self.timetable_sheet.cell(row=row, column=lecturer_col).value
            room = self.timetable_sheet.cell(row=row, column=room_col).value
            weekday = (row - 4) // 12
            is_week_odd = (row - 4) % 2 == 0
            time = ((row - 4) % 12) // 2

            weeks = filter(
                lambda week_i: first_academic_day_index <= (week_i - 1) * 6 + weekday < last_academic_day_index,
                list(range(1 if is_week_odd else 2, week_count + 2, 2)))

            if discipline.startswith("кр"):
                discipline_match = \
                    re.fullmatch(TimetableParser.PATTERN_EXCEPT_WEEKS, discipline)
                discipline = discipline_match.group(2)
                weeks = filter(
                    lambda week_i: week_i not in discipline_match.group(1).split(","),
                    weeks)
            elif re.fullmatch(TimetableParser.PATTERN_SPECIFIC_WEEKS, discipline):
                discipline_match = \
                    re.fullmatch(TimetableParser.PATTERN_SPECIFIC_WEEKS, discipline)
                discipline = discipline_match.group(2)
                weeks = [int(n) for n in discipline_match.group(1).split(",")]

            lesson = TimetableParser.get_lesson_by_type(tp, discipline, room, lecturer)

            for week in weeks:
                self.timetable[week-1][weekday][time] = lesson

    def init_timetable(self):
        week_count = utils.get_week_count()

        self.timetable = [[[NoLesson.get_instance() for _ in range(6)]
                           for _ in range(6)]
                          for _ in range(week_count+1)]

    @staticmethod
    def get_lesson_by_type(tp, discipline, room, lecturer):
        lesson = None
        if re.fullmatch(TimetableParser.PATTERN_PRACTICE, tp, flags=re.I):
            lesson = Practice(discipline, room, lecturer)
        elif re.fullmatch(TimetableParser.PATTERN_LECTURE, tp, flags=re.I):
            lesson = Lecture(discipline, room, lecturer)
        elif re.fullmatch(TimetableParser.PATTERN_LAB, tp, flags=re.I):
            lesson = Lab(discipline, room, lecturer)

        return lesson

    def add_non_academic_days(self):
        first_non_academic_weekday = 0
        last_non_academic_weekday = 5

        first_academic_weekday = utils.get_first_academic_day().weekday()
        last_academic_weekday = first_academic_weekday

        for day_i in range(first_non_academic_weekday, first_academic_weekday):
            for lesson_i in range(6):
                self.timetable[0][day_i][lesson_i] = NonAcademic.get_instance()

        for day_i in range(last_academic_weekday, last_non_academic_weekday + 1):
            for lesson_i in range(6):
                self.timetable[-1][day_i][lesson_i] = NonAcademic.get_instance()
